import openpyxl
import csv
import datetime
import win32com.client
from babel.numbers import format_decimal

#refresh excel files function
def refresh_cc_sheets(path,input_file,countries):
# Open Excel
    
    Application = win32com.client.Dispatch("Excel.Application")
 
 # Show Excel. While this is not required, it can help with debugging
    Application.Visible = 0
    Application.DisplayAlerts=False
    Application.AskToUpdateLinks = False

    for  country in countries.values():
 # Open Your Workbook
        Workbook = Application.Workbooks.open(path + input_file + country + '.xlsx')
        try:
            Workbook.UpdateLink(Name=Workbook.LinkSources())

        except Exception as e:
            print(e)
        # Refesh All
        Workbook.RefreshAll()
        Application.CalculateUntilAsyncQueriesDone()
        print(country + ' - Done')
    # Saves the Workbook
        Workbook.Save()
        Workbook.Close()
    Application.Visible = 0
    Application.DisplayAlerts=True
    Application.AskToUpdateLinks = True
 # Closes Excel
    Application.Quit()


#main part
now=datetime.datetime.now()
cur_date=now.strftime("%Y%m%d")
#dictionary for iterating between files 
countries={2:"UK", 9:"Austria", 3:"Europe", 6:"France", 1:"Germany", 11:"Italy", 13:"Spain", 7:"Switzerland",  12:"USA", 14:"Netherlands", 10:"Belgium"}
path='Z:\\800-Management\\830-Controlling\\833-Marketing\\Channel Controlling 2018\\'
input_file='Channel Controlling 2018 '
print('Started at: {}'.format(now.strftime("%H:%M")))

#refresh_cc_sheets(path,input_file,countries)


with open(path + 'DailyCostExtraction'+ '.csv', 'w', newline='') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',',
                                    quotechar='|', quoting=csv.QUOTE_MINIMAL)
            filewriter.writerow(['AffiliateGroup', 'Date', 'Cost', 'CountryId'])

for c_id, country in countries.items():


    #open the workbook
    started=datetime.datetime.now().strftime("%H:%M")
    wb = openpyxl.load_workbook(path + input_file + country + '.xlsx',read_only=True, data_only=True)

    sheets =  wb.sheetnames #list of sheet names

    #removing summary, beispiel and data pivot sheets
    i=0
    while i <=1:
        sheets.pop(0)
        i+=1
    sheets.pop()
    #sheets=list(filter(lambda sheet:sheet in ['Summary','Beispiel_1','data pivot'],sheets))
    
    #iterating between the sheets and extracting the data
    for sheet in sheets:
        #preparing lists for data that will be extracted from each sheet
        dates=[]
        costs=[]
        aff_group=[]
        country_id=[]
        data=[]
        sheet=wb[sheet]
        sheet_name=sheet[502][0].value
        # extracting only cost and date
        for row in sheet.iter_rows(min_row=3,max_row=408, min_col=1, max_col=7):
            datum = row[0].value
            cost = row[6].value
            #removing blank cells
            if datum == None:
                continue
            if cost not in [None,0,'#N/A','#VALUE!','#REF!']:
                #adding data to list
                dates.append(datum.strftime("%Y-%m-%d"))
                costs.append(format_decimal(cost, locale='de_DE')) #formating numbers in German style
                aff_group.append(sheet_name)
                country_id.append(c_id)
        #putting lists together into a list of tuples 
        data=list(zip(aff_group,dates,costs,country_id))
       
            #wiriting into the csv file
        with open(path + 'DailyCostExtraction'+ '.csv', 'a+', newline='') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',',
                                    quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for value in data:
                filewriter.writerow(value)
        #log also the sheet names for debugging        
        #with open(path + 'log'+ '.csv', 'a+', newline='') as csvfile:
        #    filewriter = csv.writer(csvfile, delimiter=',',
         #                                       quotechar='|', quoting=csv.QUOTE_MINIMAL)
         #   filewriter.writerow([sheet_name,'Done'])

    print('{}{} Done Started at:{} Ended at:{}'.format(input_file,country,started,datetime.datetime.now().strftime("%H:%M")))


    wb.close
now=datetime.datetime.now()
print('Ended at: {}'.format(now.strftime("%H:%M")))
